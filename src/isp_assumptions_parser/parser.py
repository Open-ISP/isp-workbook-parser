import os
import glob
from pathlib import Path
import yaml

import pandas as pd
import openpyxl

from .config_model import load_yaml
from .read_table import read_table


class Parser:
    """Extracts ISP inputs and assumptions data from the IASR workbbook.

    The primary intend usage is to extract all the tables and save as csv files using the save_tables method, but
    inidividual tables can be extracted as pd.DataFrames using get_table, and tables without pre-defined config can
    be extracted with user specified config using get_table_from_config. For a list of tables with pre-defined config
    compatible the version of the workbook provided use get_table_names.

    Examples

    Create a Parser instance for a particular workbook (will also check config is available for workbook version).

    >>> workbook = Parser("workbooks/5.2/2023 IASR Assumptions Workbook.xlsx")

    Save all the tables with available config to the directory example_output as csv files.

    >>> workbook.save_tables('example_output')
    """

    def __init__(
        self,
        file_path: str | Path,
        user_config_directory_path: str | Path = None
    ):
        self.file_path = self._make_path_object(file_path)
        self.file = pd.ExcelFile(self.file_path)
        self.openpyxl_file = openpyxl.load_workbook(self.file_path)
        self.workbook_version = self._get_version()
        self.default_config_path = Path(__file__).parent / Path("../../config/")
        self.config_path = self._determine_config_path(user_config_directory_path)
        self.table_configs = self._load_config()
        self.table_names = list(self.table_configs.keys())
        self.check_headers = True

    @staticmethod
    def _make_path_object(path: str | Path):
        """If the path has been provided as a string convert it to a pathlib Path object.
        """
        if not isinstance(path, Path):
            path = Path(path)
        return path

    def _get_version(self):
        """Extract the version number of the workbook from the sheet 'Change Log'.

        In the change log the version number is last value in the 'B' column. This method iterates through the values
        in the 'B' and returns the last value in the column.
        """
        sheet = self.openpyxl_file["Change Log"]
        last_value = None
        for cell in sheet["B"]:
            if cell.value is not None:
                last_value = cell.value
        return str(last_value)

    def _determine_config_path(
            self,
            user_config_directory_path: str | Path = None,
    ):
        """Determine the path to where the directory containing config yaml files are stored.

        If the user has provided a path to a directory containing the config then set this as the path to load the
        config yaml files from, otherwise use the default path and the workbook version number to construct a path
        to correct config files for the workbook version, which are shipped with the package.
        """
        if user_config_directory_path is not None:
            config_path = self._make_path_object(user_config_directory_path)
        else:
            config_path = self.default_config_path
            self._check_version_is_supported(config_path)
            config_path = config_path / Path(f"{self.workbook_version}/")
        return config_path

    def _check_version_is_supported(self, config_path):
        """Check the default config directory contains a subdirectory that matches the workbook version number.
        """
        versions = os.listdir(config_path)
        if self.workbook_version not in versions:
            raise ValueError(
                f"The workbook version {self.workbook_version} is not supported."
            )

    def _load_config(self):
        """Load all the yaml files stored in the config directory into a python dictionary with table names as keys.
        """
        pattern = os.path.join(self.config_path, '*.yaml')
        config_files = glob.glob(pattern)
        config = {}
        for file in config_files:
            config.update(load_yaml(Path(file)))
        return config

    def _check_data_ends_where_expected(
        self, tab: str, end_row: int, range: str, name: str
    ):
        """Check that the cell after the last row of the table in the second column is blank.

        While there are often notes on the data in the first cell after the first column ends, the first cell after the
        second column ends appears to be always blank. Therefore, checking that this cell is blank can be used to verify
        that the config has not specified a table end row that is before the actual last row of the table.
        """
        first_column = range.split(":")[0]
        first_col_index = openpyxl.utils.column_index_from_string(first_column)
        second_col_index = first_col_index + 1
        # We check that value in the second column is blank because sometime the row after the first column will
        # contain notes on the data.
        value_in_second_column_after_last_row = (
            self.openpyxl_file[tab].cell(row=end_row + 1, column=second_col_index).value
        )
        if value_in_second_column_after_last_row is not None:
            error_message = f"There is data in the row after the defined table end for table {name}."
            raise TableConfigError(error_message)

    @staticmethod
    def _check_for_over_run_into_another_table(data: pd.DataFrame, name: str):
        """Check that the first column of the table does not contain NA values.

        The first column of a table appears to always be an ID column with no blank values. The only reason that NA
        values would appear in this column is if the end row specified in the config occurs after the end of the
        table and inside a following table or set of notes. Therefore, the presence of NA values in the first column
        can be used to check if the end row is incorrectly specified in the config.
        """
        if data[data.columns[0]].isna().any():
            error_message = (
                f"The first column of the table {name} contains na values indicating the table end "
                f"row is incorrectly specified."
            )
            raise TableConfigError(error_message)

    @staticmethod
    def _check_for_over_run_into_notes(data: pd.DataFrame, name: str):
        """Check that the values in the first column don't contain substrings: "Notes:", "Note:", "Source:", "Sources:".

        Often the first cell after the end of the first column contains notes on the table, which appear to always
        contain one of the substrings "Notes:", "Note:", "Source:", "Sources:". If the end row is incorrectly specified
        these notes can be read into the table without creating any NA values. Therefore, checking if these substrings
        are present in any of the values in the first column is helpful in detecting if the end row is incorrectly
        specified.
        """
        notes_sub_strings = ["Notes:", "Note:", "Source:", "Sources:"]
        for sub_string in notes_sub_strings:
            if data[data.columns[0]].str.contains(sub_string).any():
                error_message = f"The first column of the table {name} contains the sub string '{sub_string}'."
                raise TableConfigError(error_message)

    @staticmethod
    def _check_no_columns_are_empty(data: pd.DataFrame, name: str):
        """Check that no columns in the table are empty.

        If the column range in the table config is incorrectly specified and the end column occurs after the end of the
        table then empty columns of data could be read into the table. Checking that no columns in the table are
        empty helps detect if the config is incorrect.
        """
        for column in data.columns:
            if data[column].isna().all():
                error_message = f"The last column of the table {name} is empty."
                raise TableConfigError(error_message)

    def _check_for_missed_column_on_right_hand_side_of_table(
        self, sheet_name: str, start_row: int, end_row: int, range: str, name: str
    ):
        """Checks if there is data in the column adjacent to last column specified in the config.

        It appears that the column adjacent to the last column in a table is always blank. Therefore, checking if
        there is data in the adjacent column can help detect when the column range in the config has been incorrectly
        specified.
        """
        last_column = range.split(":")[1]
        last_col_index = openpyxl.utils.column_index_from_string(last_column)
        column_next_to_last_column = openpyxl.utils.get_column_letter(
            last_col_index + 1
        )
        column_next_to_last_column = (
            column_next_to_last_column + ":" + column_next_to_last_column
        )
        range_error = False
        try:
            data = pd.read_excel(
                self.file,
                sheet_name=sheet_name,
                header=start_row,
                usecols=column_next_to_last_column,
                nrows=(end_row - start_row),
            )
            if not data[data.columns[0]].isna().all():
                range_error = True
        except pd.errors.ParserError:
            range_error = False

        if range_error:
            error_message = f"There is data in the column adjacent to the last column in the table {name}."
            raise TableConfigError(error_message)

    @staticmethod
    def _check_headers(data: pd.DataFrame, table_config):
        """Check that the column names as read from the workbook match the expected column names from the config.

        This check exists to catch instance where column names have been changed, a column has been deleted, or a column
        has been added. By checking the workbook against a set of expected names
        """
        if list(data.columns) != table_config.header_names:
            name = table_config.name
            error_message = f"Column names for the table {name} don't match the header names provided in the config."
            raise TableConfigError(error_message)

    def _check_table(self, data, table_config):
        if isinstance(table_config.header_rows, list):
            start_row = table_config.header_rows[0]
        else:
            start_row = table_config.header_rows

        self._check_data_ends_where_expected(
            table_config.sheet_name,
            table_config.end_row,
            table_config.column_range,
            table_config.name,
        )
        self._check_for_missed_column_on_right_hand_side_of_table(
            table_config.sheet_name,
            start_row,
            table_config.end_row,
            table_config.column_range,
            table_config.name,
        )
        self._check_for_over_run_into_another_table(data, table_config.name)
        self._check_for_over_run_into_notes(data, table_config.name)
        self._check_no_columns_are_empty(data, table_config.name)
        if self.check_headers:
            self._check_headers(data, table_config)

    def get_table_names(self) -> list[str]:
        """Returns the list of tables that there is config for extracting from the workbook version provided.

        Examples:
        >>> workbook = Parser("workbooks/5.2/2023 IASR Assumptions Workbook.xlsx")

        >>> workbook.get_table_names()

        Returns:
            List of the tables that there is configuration information for extracting from the workbook.
        """
        return self.table_names

    def get_table_from_config(
        self, table_config, config_checks: bool = True
    ) -> pd.DataFrame:
        """Retrieves a table from the assumptions workbook using the config provided and returns as pd.DataFrame.

        Examples:

        >>> from isp_assumptions_parser import Table

        >>> config = Table(
        ... name='existing_generators_summary',
        ... sheet_name='Summary Mapping',
        ... header_rows=[4, 5, 6],
        ... end_row=256,
        ... column_range="B:AC",
        ... )

        >>> workbook = Parser("workbooks/5.2/2023 IASR Assumptions Workbook.xlsx")

        >>> workbook.get_table_from_config(config)

        Args:
            table_config:
            config_checks: A boolean that specifies whether to check the table config by checking if the data
                starts and ends where expected and the workbook header matches the config header.

        """
        data = read_table(self.file, table_config)
        if config_checks:
            self._check_table(data, table_config)
        return data

    def get_table(self, table_name: str, config_checks: bool = True) -> pd.DataFrame:
        """Retrieves a table from the assumptions workbook and returns as pd.DataFrame.

        Examples
        >>> workbook = Parser("workbooks/5.2/2023 IASR Assumptions Workbook.xlsx")

        >>> workbook.get_table('existing_generators_summary')
                            Generator  ... Auxiliary load (%)
        2                   Bayswater  ...     Black Coal NSW
        3                     Eraring  ...     Black Coal NSW
        4                    Mt Piper  ...     Black Coal NSW
        5               Vales Point B  ...     Black Coal NSW
        6                   Callide B  ...     Black Coal QLD
        ..                        ...  ...                ...
        247  Stockyard Hill Wind Farm  ...               Wind
        248          Waubra Wind Farm  ...               Wind
        249    Yaloak South Wind Farm  ...               Wind
        250          Yambuk Wind Farm  ...               Wind
        251          Yendon Wind Farm  ...               Wind
        <BLANKLINE>
        [250 rows x 27 columns]

        Args:
            table_name: string specifying the table to retrieve.
            config_checks: A boolean that specifies whether to check the tabe config by checking if the data
                starts and ends where expected and the workbook header matches the config header.
        """
        if not isinstance(table_name, str):
            ValueError("The parameter table_name must be provided as a string.")

        if table_name not in self.table_names:
            ValueError(
                "The table_name provided is not in the config for this workbook version."
            )

        table_config = self.table_configs[table_name]
        data = self.get_table_from_config(table_config, config_checks=config_checks)
        return data

    def save_tables(
        self,
        directory: str | Path,
        tables: list[str] | str = "all",
        config_checks: bool = True,
    ) -> None:
        """Saves tables from the assumptions workbook to the specified directory as CSV files.

        Examples:
        >>> workbook = Parser("workbooks/5.2/2023 IASR Assumptions Workbook.xlsx")

        >>> workbook.save_tables(
        ... directory="example_output",
        ... tables=['existing_generators_summary']
        ... )

        Args:
            tables: A list of strings specifying which tables to extract from the workbook and save, or the
                str 'all', which will result in all the tables the isp-assumptions-parser has config for being
                saved.
            directory: A str specifying the path to the directory or a pathlib Path object.
            config_checks: A boolean that specifies whether to check the tabe config by checking if the data
                starts and ends where expected and the workbook header matches the config header.

        Returns:
            None
        """
        directory = self._make_path_object(directory)

        if not directory.is_dir():
            ValueError("The path provided is not a directory.")

        if not (isinstance(tables, str) or isinstance(tables, list)):
            ValueError("The parameter tables must be provided as str or list[str].")

        if isinstance(tables, str) and tables != "all":
            ValueError(
                "If the parameter tables is provided as a str it must \n",
                f"have the value 'all' but '{tables}' was provided.",
            )

        if tables == "all":
            tables = self.get_table_names()

        for table_name in tables:
            table = self.get_table(table_name, config_checks=config_checks)
            save_path = directory / Path(f"{table_name}.csv")
            table.to_csv(save_path)

    def save_config_with_headers(
        self,
        directory: str | Path,
        tables: list[str] | str = "all",
        config_checks: bool = True,
    ) -> None:
        """Saves a new version of the config with the header names as read from Excel Workbook.

        The purpose of this method is to allow the user to write a new config without manually specifying the table
        headers. Then this method can be run and a version of the config save with the header names automatically
        saved to a new copy of the config files. To do this, tables are read without checking column names against
        the header names in the config. Note, the purpose of normally checking the column names against the config
        is to make sure the data hasn't changed format since the config for the version was written. Effectively, this
        method can be used to create a config with a snapshot of the headers for a given version of the workbook.

        Examples:
        >>> workbook = Parser("workbooks/5.2/2023 IASR Assumptions Workbook.xlsx")

        >>> workbook.save_config_with_headers(
        ... directory="example_output/config",
        ... )

        Args:
            tables: A list of strings specifying which tables to extract from the workbook and save, or the
                str 'all', which will result in all the tables the isp-assumptions-parser has config for being
                saved.
            directory: A str specifying the path to the directory or a pathlib Path object where the config files
                will be saved.
            config_checks: A boolean that specifies whether to check the tabe config by checking if the data
                starts and ends where expected and the workbook header matches the config header. Note checking
                the header names is always disabled for this method.

        Returns:
            None
        """

        directory = self._make_path_object(directory)

        if not directory.is_dir():
            ValueError("The path provided is not a directory.")

        if not (isinstance(tables, str) or isinstance(tables, list)):
            ValueError("The parameter tables must be provided as str or list[str].")

        if isinstance(tables, str) and tables != "all":
            ValueError(
                "If the parameter tables is provided as a str it must \n",
                f"have the value 'all' but '{tables}' was provided.",
            )

        if tables == "all":
            tables = self.get_table_names()

        tables_config_by_sheet = {}

        field_save_order = [
            "sheet_name",
            "header_rows",
            "end_row",
            "column_range",
            "header_names",
        ]

        self.check_headers = False

        for table_name in tables:
            table_config = self.table_configs[table_name]
            table = self.get_table(table_name, config_checks=config_checks)
            table_config.header_names = list(table.columns)
            if table_config.sheet_name not in tables_config_by_sheet:
                tables_config_by_sheet[table_config.sheet_name] = {}
            table_config = table_config.dict()
            del table_config["name"]
            table_config = {
                key: table_config[key]
                for key in field_save_order
                if key in table_config
            }
            tables_config_by_sheet[table_config["sheet_name"]][table_name] = (
                table_config
            )

        self.check_headers = True

        for sheet_name, tables in tables_config_by_sheet.items():
            file_name = sheet_name.lower().rstrip().replace(" ", "_") + ".yaml"
            file_path = directory / Path(file_name)
            with open(file_path, "w") as f:
                yaml.safe_dump(tables, f, default_flow_style=False, sort_keys=False)
                f.close()


class TableConfigError(Exception):
    """Raise for table configuration failing check."""
