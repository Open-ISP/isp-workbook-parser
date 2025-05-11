import glob
import os
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import openpyxl.utils
import pandas as pd
from thefuzz import process

from .config_model import TableConfig, load_yaml
from .read_table import _find_data_column_index, read_table
from .sanitisers import _values_casting_and_sanitisation


class Parser:
    """Extracts ISP inputs and assumptions data from the IASR workbbook.

    If a directory path containing configs is provided by the user, this is used as the
    path to load the config YAM files from. Otherwise, a default path is constructed
    using the workbook version number. This uses config files shipped with the package.

    For a list of tables with pre-defined config (i.e. those shipped with the package)
    compatible with the version of the workbook provided, use `Parser.get_table_names`.

    Usage:
    1. Extract all the tables specified in the `Parser.config_path`, or, if this is not
        provided, in the `Parser.default_config_path`. Saves these as csv files using the
        `save_tables` method
    2. Extract individual tables as `pd.DataFrame`s using `Parser.get_table`. Tables
        specified should be included in the pre-defined config for the provided
        workbook's version.
    3. Extract tables using a user-specified config with `Parser.get_table_from_config`.

    Examples:

    Create a Parser instance for a particular workbook. Will also check config is available for workbook version.

    >>> workbook = Parser("workbooks/6.0/2024-isp-inputs-and-assumptions-workbook.xlsx") # doctest: +SKIP

    Save all the tables with available config to the directory example_output as csv files.

    >>> workbook.save_tables('example_output') # doctest: +SKIP
    """

    def __init__(
        self, file_path: str | Path, user_config_directory_path: str | Path = None
    ) -> None:
        self.file_path = self._make_path_object(file_path)
        self.file = pd.ExcelFile(self.file_path)
        self.openpyxl_file = openpyxl.load_workbook(self.file_path)
        self.workbook_version = self._get_version()
        self.default_config_path = Path(__file__).parent.parent / Path(
            "isp_table_configs"
        )
        self.config_path = self._determine_config_path(user_config_directory_path)
        self.table_configs = self._load_config()
        self.table_names_by_sheet = self._get_table_names_by_sheet()

    @staticmethod
    def _make_path_object(path: str | Path) -> Path:
        """If the path has been provided as a string convert it to a pathlib Path object."""
        if not isinstance(path, Path):
            path = Path(path)
        return path

    def _get_version(self) -> str:
        """Extract the version number of the workbook from the sheet 'Change Log'.

        In the change log the version number is last value in the 'B' column. This method iterates through the values
        in the 'B' and returns the last value in the column.
        """
        sheet = self.openpyxl_file["Change Log"]
        last_value = None
        for cell in sheet["B"]:
            if cell.value is not None:
                last_value = cell.value
        version = float(last_value)
        return str(version)

    def _determine_config_path(
        self,
        user_config_directory_path: str | Path = None,
    ) -> Path:
        """Determine the path to where the directory containing config YAML files are stored.

        If the user has provided a path to a directory containing the config then set this as the path to load the
        config YAML files from, otherwise use the default path and the workbook version number to construct a path
        to correct config files for the workbook version, which are shipped with the package.
        """
        if user_config_directory_path is not None:
            config_path = self._make_path_object(user_config_directory_path)
        else:
            config_path = self.default_config_path
            self._check_version_is_supported(config_path)
            config_path = config_path / Path(f"{self.workbook_version}/")
        return config_path

    def _check_version_is_supported(self, config_path) -> None:
        """Check the default config directory contains a subdirectory that matches the workbook version number."""
        versions = os.listdir(config_path)
        if self.workbook_version not in versions:
            raise ValueError(
                f"The workbook version {self.workbook_version} is not supported."
            )

    def _load_config(self) -> dict[str, dict[str, Any]]:
        """Load all the YAML files stored in the config directory into a nested dictionary with sheet names as keys
        and table names as second level keys. For robustness across workbook versions, the config sheet name
        is matched with a workbook sheet name in case-agnostic manner.
        """
        pattern = os.path.join(self.config_path, "*.yaml")
        config_files = glob.glob(pattern)
        configs = {}
        for file in config_files:
            config_dict = load_yaml(Path(file))
            for config_name in config_dict.keys():
                config = config_dict[config_name]
                config_sheet_name_lowercase = config.sheet_name.lower()
                sheet_names = [
                    sheet_name
                    for sheet_name in self.file.sheet_names
                    if sheet_name.lower() == config_sheet_name_lowercase
                ]
                if len(sheet_names) > 1:
                    raise TableConfigError(
                        f"Workbook sheet '{config.sheet_name}' is not unique"
                    )
                elif len(sheet_names) < 1:
                    raise TableConfigError(
                        f" Sheet '{config.sheet_name}' cannot be found in the workbook"
                    )
                else:
                    config.sheet_name = sheet_names.pop()
                config_dict[config_name] = config
            configs.update(config_dict)
        return configs

    def _get_table_names_by_sheet(self):
        table_names_by_sheet = {}
        for table_name, config in self.table_configs.items():
            if config.sheet_name not in table_names_by_sheet:
                table_names_by_sheet[config.sheet_name] = []
            table_names_by_sheet[config.sheet_name].append(table_name)
        sorted_table_names_by_sheet = dict(sorted(table_names_by_sheet.items()))
        for sheet_name, tables in sorted_table_names_by_sheet.items():
            sorted_table_names_by_sheet[sheet_name] = sorted(tables)
        return sorted_table_names_by_sheet

    def _check_data_ends_where_expected(
        self, tab: str, end_row: int, range: str, name: str
    ) -> None:
        """Check that the cell after the last row of the table in the second column is blank.

        While there are often notes on the data in the first cell after the first column ends, the first cell after the
        second column ends appears to be always blank. Therefore, checking that this cell is blank can be used to verify
        that the config has not specified a table end row that is before the actual last row of the table.
        """
        first_column = range.split(":")[0]
        first_col_index = openpyxl.utils.column_index_from_string(first_column)
        second_col_index = first_col_index + 1
        # We check that value in the second column is blank because sometime the row after the first column will
        # contain notes on the data.
        value_in_second_column_after_last_row = (
            self.openpyxl_file[tab].cell(row=end_row + 1, column=second_col_index).value
        )
        if (
            value_in_second_column_after_last_row is not None
            and value_in_second_column_after_last_row not in ["", " ", "\u00a0"]
        ):
            error_message = f"There is data in the row after the defined table end for table {name}."
            raise TableConfigError(error_message)

    def _check_no_data_above_first_header_row(
        self, tab: str, header_rows: int, range: str, name: str
    ) -> None:
        """Check that the cell before the first header row of the table in the second column is blank.

        While there are often notes on the data in the first cell above the first column, the first cell above the
        second column appears to be always blank. Therefore, checking that this cell is blank can be used to verify
        that the config has not specified a table header row that is after the first header row of the table.
        """
        first_column = range.split(":")[0]
        first_col_index = openpyxl.utils.column_index_from_string(first_column)
        second_col_index = first_col_index + 1

        if isinstance(header_rows, int):
            first_header_row = header_rows
        else:
            first_header_row = header_rows[0]

        # We check that value in the second column is blank because sometime the row after the first column will
        # contain notes on the data.
        value_in_second_column_after_last_row = (
            self.openpyxl_file[tab]
            .cell(row=first_header_row - 1, column=second_col_index)
            .value
        )
        if (
            value_in_second_column_after_last_row is not None
            and value_in_second_column_after_last_row not in ["", " ", "\u00a0"]
        ):
            error_message = f"There is data or a header above the first header row for table {name}."
            raise TableConfigError(error_message)

    @staticmethod
    def _check_for_over_run_into_another_table(data: pd.DataFrame, name: str) -> None:
        """Check that the first column of the table does not contain NA values.

        The first column of a table appears to always be an ID column with no blank values. The only reason that NA
        values would appear in this column is if the end row specified in the config occurs after the end of the
        table and inside a following table or set of notes. Therefore, the presence of NA values in the first column
        can be used to check if the end row is incorrectly specified in the config.
        """
        if data[data.columns[0]].isna().any():
            error_message = (
                f"The first column of the table {name} contains na values indicating the table end "
                f"row is incorrectly specified."
            )
            raise TableConfigError(error_message)

    @staticmethod
    def _check_for_over_run_into_notes(data: pd.DataFrame, name: str) -> None:
        """Check that the values in the first column don't contain substrings: "Notes:", "Note:", "Source:", "Sources:".

        Often the first cell after the end of the first column contains notes on the table, which appear to always
        contain one of the substrings "Notes:", "Note:", "Source:", "Sources:". If the end row is incorrectly specified
        these notes can be read into the table without creating any NA values. Therefore, checking if these substrings
        are present in any of the values in the first column is helpful in detecting if the end row is incorrectly
        specified.
        """
        notes_sub_strings = ["Notes:", "Note:", "Source:", "Sources:"]
        for sub_string in notes_sub_strings:
            if data[data.columns[0]].astype(str).str.contains(sub_string).any():
                error_message = f"The first column of the table {name} contains the sub string '{sub_string}'."
                raise TableConfigError(error_message)

    @staticmethod
    def _check_last_column_isnt_empty(data: pd.DataFrame, name: str) -> None:
        """Check the column in the table isnt empty.

        If the column range in the table config is incorrectly specified and the end column occurs after the end of the
        table then empty columns of data could be read into the table. Checking if the last column in the table is
        empty helps detect if the config is incorrect.
        """
        if "Unnamed" in data.columns[-1]:
            error_message = f"The last column of the table {name} is empty."
            raise TableConfigError(error_message)

    @staticmethod
    def _check_columns_unique(data: pd.DataFrame, name: str) -> None:
        """Check that columns in the data are unique

        Unique columns names are required for sanitisation to work without error (i.e. in
        `isp_workbook_parser.sanitisers._values_casting_and_sanitisation`). If an error
        is raised, this may indicate that the config is incorrect.

        """
        if len(data.columns) != len(data.columns.drop_duplicates()):
            error_message = f"There are duplicate column names in the table {name}."
            raise TableConfigError(error_message)

    def _check_for_missed_column_on_right_hand_side_of_table(
        self, sheet_name: str, start_row: int, end_row: int, range: str, name: str
    ) -> None:
        """Checks if there is data in the column adjacent to last column specified in the config.

        It appears that the column adjacent to the last column in a table is always blank. Therefore, checking if
        there is data in the adjacent column can help detect when the column range in the config has been incorrectly
        specified.
        """
        last_column = range.split(":")[1]
        last_col_index = openpyxl.utils.column_index_from_string(last_column)
        column_next_to_last_column = openpyxl.utils.get_column_letter(
            last_col_index + 1
        )
        column_next_to_last_column = (
            column_next_to_last_column + ":" + column_next_to_last_column
        )
        range_error = False
        try:
            data = pd.read_excel(
                self.file,
                sheet_name=sheet_name,
                header=start_row - 1,
                usecols=column_next_to_last_column,
                nrows=(end_row - start_row),
            )
            data = data[~data.isin(["`"])]  # explicit exceptions for messy data
            # replace non breaking space with empty string
            data = data.replace("\xa0", "")
            # replace empty strings with np.nan
            data = data.replace("", np.nan)
            if not data[data.columns[0]].isna().all():
                range_error = True
        except pd.errors.ParserError:
            range_error = False

        if range_error:
            error_message = f"There is data in the column adjacent to the last column in the table {name}."
            raise TableConfigError(error_message)

    def _check_for_missed_column_on_left_hand_side_of_table(
        self, sheet_name: str, start_row: int, end_row: int, range: str, name: str
    ) -> None:
        """Checks if there is data in the column adjacent to first column specified in the config.

        It appears that the column adjacent to the first column in a table is always blank. Therefore, checking if
        there is data in the adjacent column can help detect when the column range in the config has been incorrectly
        specified.
        """
        first_column = range.split(":")[0]
        first_col_index = openpyxl.utils.column_index_from_string(first_column)
        column_next_to_first_column = openpyxl.utils.get_column_letter(
            first_col_index - 1
        )
        column_next_to_first_column = (
            column_next_to_first_column + ":" + column_next_to_first_column
        )
        range_error = False
        try:
            data = pd.read_excel(
                self.file,
                sheet_name=sheet_name,
                header=start_row - 1,
                usecols=column_next_to_first_column,
                nrows=(end_row - start_row),
            )
            if data[data.columns[0]].isna().all():
                range_error = False
            elif "DO NOT DELETE THIS COLUMN" in data.columns[0] or first_column == "B":
                range_error = False
            else:
                range_error = True
        except pd.errors.ParserError:
            range_error = False

        if range_error:
            error_message = f"There is data in the column adjacent to the first column in the table {name}."
            raise TableConfigError(error_message)

    def _check_if_header_row_and_end_row_are_on_sheet(self, table_config) -> None:
        """Checks if first row of header and end_row are within the sheet."""
        if isinstance(table_config.header_rows, int):
            first_header_row = table_config.header_rows
        else:
            first_header_row = table_config.header_rows[0]

        if first_header_row > self.openpyxl_file[table_config.sheet_name].max_row:
            error_message = f"The first header row for table {table_config.name} is not within the excel sheet."
            raise TableConfigError(error_message)
        if table_config.end_row > self.openpyxl_file[table_config.sheet_name].max_row:
            error_message = f"The end_row for table {table_config.name} is not within the excel sheet."
            raise TableConfigError(error_message)

    def _check_if_start_and_end_column_are_on_sheet(self, table_config) -> None:
        """Checks if first column and last column in config are within the sheet."""
        first_column = table_config.column_range.split(":")[0]
        first_col_index = openpyxl.utils.column_index_from_string(first_column)
        if first_col_index > self.openpyxl_file[table_config.sheet_name].max_column:
            error_message = f"The first column for table {table_config.name} is not within the excel sheet."
            raise TableConfigError(error_message)

        last_column = table_config.column_range.split(":")[1]
        last_col_index = openpyxl.utils.column_index_from_string(last_column)
        if last_col_index > self.openpyxl_file[table_config.sheet_name].max_column:
            error_message = f"The last column for table {table_config.name} is not within the excel sheet."
            raise TableConfigError(error_message)

    def _check_table(self, data, table_config) -> None:
        if isinstance(table_config.header_rows, list):
            start_row = table_config.header_rows[0]
            last_header_row = table_config.header_rows[-1]
        else:
            start_row = table_config.header_rows
            last_header_row = table_config.header_rows
        self._check_no_data_above_first_header_row(
            table_config.sheet_name,
            table_config.header_rows,
            table_config.column_range,
            table_config.name,
        )
        self._check_data_ends_where_expected(
            table_config.sheet_name,
            table_config.end_row,
            table_config.column_range,
            table_config.name,
        )
        self._check_for_missed_column_on_right_hand_side_of_table(
            table_config.sheet_name,
            last_header_row,
            table_config.end_row,
            table_config.column_range,
            table_config.name,
        )
        self._check_for_missed_column_on_left_hand_side_of_table(
            table_config.sheet_name,
            start_row,
            table_config.end_row,
            table_config.column_range,
            table_config.name,
        )
        self._check_last_column_isnt_empty(data, table_config.name)
        self._check_for_over_run_into_another_table(data, table_config.name)
        self._check_for_over_run_into_notes(data, table_config.name)

    def _postprocess_percentage_columns_between_0_and_100(
        self, data: pd.DataFrame, table_config: TableConfig
    ) -> pd.DataFrame:
        """Multiplies cells with percentage formatting by 100 to ensure that
        percentage values are between 0 and 100.

        `pandas.read_excel` parses percentage formatted data as values between 0 and 1.
        This postprocssor uses `openpyxl` to check the number format of each cell, and
        then multiplies any percentage formatted cells by 100 to ensure that all
        percentage data is a value between 0 and 100 (N.B. some percentage data are
        formatted as integers between 0 and 100)

        Args:
            data: `pandas.DataFrame` produced by `read_table`
            table_config: `TableConfig` for the data table

        Returns:
            `pandas.DataFrame` with percentage columns multiplied by 100 (i.e.
            values should be between 0 and 100)
        """
        percentage_columns = []
        sheet = self.openpyxl_file[table_config.sheet_name]
        min_col, max_col = [
            openpyxl.utils.column_index_from_string(col_alphabetical)
            for col_alphabetical in table_config.column_range.split(":")
        ]
        if isinstance(table_config.header_rows, list):
            min_row = table_config.header_rows[-1] + 1
        else:
            min_row = table_config.header_rows + 1
        for col in sheet.iter_cols(
            min_row=min_row,
            max_row=table_config.end_row,
            min_col=min_col,
            max_col=max_col,
        ):
            percentage_cells = []
            skipped_rows = 0
            for cell in col:
                if sr := table_config.skip_rows:
                    if isinstance(sr, list) and cell.row in sr:
                        skipped_rows += 1
                        continue
                    elif isinstance(sr, int) and cell.row == sr:
                        skipped_rows += 1
                        continue
                if isinstance(cell.value, (int, float)) and "%" in cell.number_format:
                    percentage_cells.append(
                        (
                            cell.row - min_row - skipped_rows,
                            _find_data_column_index(
                                cell.column_letter, table_config.column_range
                            ),
                        )
                    )

            # add the data column index if the entire column consists of percentage values
            # else, add the individual cells as a list of tuples
            if len(percentage_cells) == (table_config.end_row - min_row + 1):
                percentage_columns.append(set(x[1] for x in percentage_cells).pop())
            else:
                percentage_columns.append(percentage_cells)

        for col in percentage_columns:
            if isinstance(col, int):
                data.iloc[:, col] *= 100
            elif isinstance(col, list):
                for cell in col:
                    data.iloc[cell[0], cell[1]] *= 100
        return data

    def get_table_names(self) -> list[str]:
        """Returns a dict of table names by sheet name that there is config for.

        Examples:
        >>> workbook = Parser("workbooks/6.0/2024-isp-inputs-and-assumptions-workbook.xlsx")

        >>> names = workbook.get_table_names()

        >>> names['Build limits']
        ['initial_build_limits', 'rez_transmission_modifiers']


        Returns:
            List of the tables that there is configuration information for extracting from the workbook.
        """
        return self.table_names_by_sheet

    def get_table_from_config(
        self, table_config: TableConfig, config_checks: bool = True
    ) -> pd.DataFrame:
        """Retrieves a table from the assumptions workbook using the config provided and returns as pd.DataFrame.

        Examples:

        >>> import pandas as pd
        >>> from isp_workbook_parser import TableConfig

        >>> pd.set_option("display.max_columns", 4)
        >>> pd.set_option("display.width", None)

        >>> config = TableConfig(
        ... name='existing_generators_summary',
        ... sheet_name='Summary Mapping',
        ... header_rows=[4, 5, 6],
        ... end_row=258,
        ... column_range="B:AB",
        ... )

        >>> workbook = Parser("workbooks/6.0/2024-isp-inputs-and-assumptions-workbook.xlsx")

        >>> workbook.get_table_from_config(config).head()
          Existing generator     Technology type  ...            MLF Auxiliary load (%)
        0          Bayswater  Steam Sub Critical  ...      Bayswater     Black Coal NSW
        1            Eraring  Steam Sub Critical  ...        Eraring     Black Coal NSW
        2           Mt Piper  Steam Sub Critical  ...       Mt Piper     Black Coal NSW
        3      Vales Point B  Steam Sub Critical  ...  Vales Point B     Black Coal NSW
        4          Callide B  Steam Sub Critical  ...      Callide B     Black Coal QLD
        <BLANKLINE>
        [5 rows x 27 columns]

        Args:
            table_config: A table configuration.
            config_checks: Specifies whether to check the table config by checking if the data
                starts and ends where expected and the workbook header matches the config header.

        """
        if config_checks:
            self._check_if_header_row_and_end_row_are_on_sheet(table_config)
            self._check_if_start_and_end_column_are_on_sheet(table_config)
        data = read_table(self.file, table_config)
        self._check_columns_unique(data, table_config.name)
        data = _values_casting_and_sanitisation(data)
        data = self._postprocess_percentage_columns_between_0_and_100(
            data, table_config
        )
        if config_checks:
            self._check_table(data, table_config)
        return data

    def get_table(self, table_name: str, config_checks: bool = True) -> pd.DataFrame:
        """Retrieves a table from the assumptions workbook and returns as `pd.DataFrame`.

        Examples
        >>> workbook = Parser("workbooks/6.0/2024-isp-inputs-and-assumptions-workbook.xlsx")

        >>> workbook.get_table('wind_high_capacity_factors').head()
          Wind High_REZ ID  ... Wind High_Avg of ref years
        0               Q1  ...                   0.456914
        1               Q2  ...                   0.417142
        2               Q3  ...     Resource limit of 0 MW
        3               Q4  ...                    0.34437
        4               Q5  ...                   0.326029
        <BLANKLINE>
        [5 rows x 17 columns]

        Args:
            table_name: Specified the table to retrieve.
            config_checks: Specifies whether to check the tabe config by checking if the data
                starts and ends where expected and the workbook header matches the config header.
        """
        if not isinstance(table_name, str):
            raise ValueError("The parameter table_name must be provided as a string.")
        if table_name not in self.table_configs.keys():
            closest = process.extractOne(table_name, self.table_configs.keys())[0]
            raise ValueError(
                f"The table_name ({table_name}) provided is not in the config for this workbook version."
                + f" Did you mean '{closest}'?"
            )

        table_config = self.table_configs[table_name]
        data = self.get_table_from_config(table_config, config_checks=config_checks)
        return data

    def save_tables(
        self,
        directory: str | Path,
        tables: list[str] | str = "all",
        config_checks: bool = True,
    ) -> None:
        """Saves tables from the provided workbook to the specified directory as CSV files.

        Examples:
        >>> workbook = Parser("workbooks/6.0/2024-isp-inputs-and-assumptions-workbook.xlsx") # doctest: +SKIP

        >>> workbook.save_tables(directory="example_output") # doctest: +SKIP

        Args:
            tables: Which tables to extract from the workbook and save, or the str 'all',
                which will result in all the tables the isp-workbook-parser has config for being
                saved.
            directory: Path to the directory or a pathlib Path object.
            config_checks: Specifies whether to check the tabe config by checking if the data
                starts and ends where expected.

        Returns:
            None
        """
        directory = self._make_path_object(directory)
        if not directory.exists():
            directory.mkdir(parents=True)

        if not directory.is_dir():
            raise ValueError("The path provided is not a directory.")

        if not (isinstance(tables, str) or isinstance(tables, list)):
            raise ValueError(
                "The parameter tables must be provided as str or list[str]."
            )

        if isinstance(tables, str) and tables != "all":
            raise ValueError(
                "If the parameter tables is provided as a str it must \n",
                f"have the value 'all' but '{tables}' was provided.",
            )

        if tables == "all":
            tables = self.table_configs.keys()

        for table_name in tables:
            table = self.get_table(table_name, config_checks=config_checks)
            save_path = directory / Path(f"{table_name}.csv")
            table.to_csv(save_path, index=False)


class TableConfigError(Exception):
    """Raise for table configuration failing check."""
